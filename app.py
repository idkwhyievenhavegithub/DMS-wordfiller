## DMS Word Template filler
## Generate custom documentation for the roll out of Microsoft 365.
import os

from docx import Document
from docxtpl import DocxTemplate, RichText
from openpyxl import load_workbook


main_path = r"/Users/jayden/code/DMS-wordfiller/"
template_path = os.path.join(main_path, 'dms_microsoft_template.docx')
workbook_path = os.path.join(main_path, 'pw_book.xlsx')

wb = load_workbook(filename = workbook_path)
sheet_ranges = wb['Sheet1']


template = DocxTemplate(template_path)



total = sheet_ranges.max_row
for x in range(1,total):
    Row = str(x)
    Name = sheet_ranges["A"+Row].value
    Email = sheet_ranges["B"+Row].value
    PW = sheet_ranges["C"+Row].value
    to_fill_in = {"NAME": Name,
              "EMAIL": Email,
              "PASSWORD": PW
              }
    template.render(to_fill_in)
    filename = Name+" Access DMS Microsoft.docx"
    filled_path = os.path.join(main_path+"Render/", filename)
    template.save(filled_path)